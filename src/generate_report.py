"""
generate_report.py
================================================================
Run the baseline allocation and produce a set of analysis charts +
an Excel workbook with the concrete assignment.  This is the script
that answers the supervisor's request: "show concrete results, show
the output."

Outputs (in output/):
    allocation_2026.xlsx        full per-student assignment + statistics
    rank_distribution.png       how many students at rank 1/2/3
    demand_vs_capacity.png      the oversubscription picture
    partner_groups.png          partner co-assignment status
    project_utilization.png     occupancy of every opened project
================================================================
"""

import pickle
from collections import defaultdict
from pathlib import Path

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from spa_milp import load_data, solve

OUTPUT_DIR = Path(__file__).resolve().parent.parent / "output"

C_RANK = {1: "#27AE60", 2: "#2980B9", 3: "#C0392B"}
C_BLUE = "#193764"
C_RED  = "#C0392B"


def save_excel(data, res, path):
    projects = data["projects"]
    wb = Workbook()
    ws = wb.active
    ws.title = "Assignments"
    head = ["Student", "Email", "Project ID", "Project Title", "Supervisor", "Rank", "Track-match"]
    ws.append(head)
    for c in ws[1]:
        c.fill = PatternFill("solid", fgColor=C_BLUE.lstrip("#"))
        c.font = Font(color="FFFFFF", bold=True)

    track_stats = res.get("track_stats")
    for a in sorted(res["assignments"], key=lambda a: (a["rank"] or 9, a["student"]["name"])):
        s, p = a["student"], a["project"]
        if p:
            ws.append([s["name"], s["email"], p, projects[p]["name"],
                       projects[p]["supervisor"], a["rank"], ""])
        else:
            ws.append([s["name"], s["email"], "—", "NOT ASSIGNED", "", "—", ""])
    for col, w in zip("ABCDEFG", [26, 30, 16, 55, 22, 6, 12]):
        ws.column_dimensions[col].width = w

    ws2 = wb.create_sheet("Statistics")
    ws2.append(["Metric", "Value"])
    for c in ws2[1]:
        c.fill = PatternFill("solid", fgColor=C_BLUE.lstrip("#"))
        c.font = Font(color="FFFFFF", bold=True)
    rows = [
        ["Students", len(data["students"])],
        ["Assigned", sum(1 for a in res["assignments"] if a["project"])],
        ["Satisfaction %", round(res["satisfaction_pct"], 1)],
        ["Worst rank", res["worst_rank"]],
        ["Projects opened", res["n_opened"]],
        ["Partner groups satisfied", f"{res['n_partners_ok']}/{len(res['partner_status'])}"],
        ["Unstable students", len(res["unstable"])],
    ]
    for r in rows:
        ws2.append(r)
    wb.save(path)


def chart_rank(res, path):
    fig, ax = plt.subplots(figsize=(6, 4))
    ranks = [1, 2, 3]
    counts = [res["rank_counts"].get(r, 0) for r in ranks]
    bars = ax.bar(ranks, counts, color=[C_RANK[r] for r in ranks], width=0.6,
                  edgecolor="white", linewidth=1.5)
    for b, c in zip(bars, counts):
        ax.text(b.get_x()+b.get_width()/2, b.get_height()+0.4, str(c),
                ha="center", va="bottom", fontsize=13, fontweight="bold")
    ax.set_xticks(ranks); ax.set_xlabel("Assigned choice (rank)")
    ax.set_ylabel("Number of students")
    ax.set_title(f"Rank Distribution — {res['satisfaction_pct']:.1f}% satisfaction",
                 fontsize=13, fontweight="bold")
    ax.spines[["top", "right"]].set_visible(False)
    fig.tight_layout(); fig.savefig(path, dpi=150, facecolor="white"); plt.close(fig)


def chart_demand(data, path):
    projects, students = data["projects"], data["students"]
    demand = defaultdict(int)
    for s in students:
        for p in s["choices"]:
            demand[p] += 1
    items = sorted(demand.items(), key=lambda kv: kv[1], reverse=True)[:15]
    names = [projects[p]["name"][:34] for p, _ in items]
    dem = [d for _, d in items]
    cap = [projects[p]["quota"][1] for p, _ in items]
    y = range(len(items))
    fig, ax = plt.subplots(figsize=(9, 6))
    ax.barh(y, dem, height=0.4, color=C_BLUE, label="Demand")
    ax.barh([i+0.4 for i in y], cap, height=0.4, color="#BDC3C7", label="Capacity")
    ax.set_yticks([i+0.2 for i in y]); ax.set_yticklabels(names, fontsize=9)
    ax.invert_yaxis(); ax.set_xlabel("Students"); ax.legend()
    ax.set_title("Top 15 Projects — Demand vs Capacity", fontsize=13, fontweight="bold")
    ax.spines[["top", "right"]].set_visible(False)
    fig.tight_layout(); fig.savefig(path, dpi=150, facecolor="white"); plt.close(fig)


def chart_partners(res, path):
    ps = res["partner_status"]
    fig, ax = plt.subplots(figsize=(8, 0.5*len(ps)+1))
    y = range(len(ps))
    colors = ["#27AE60" if p["satisfied"] else C_RED for p in ps]
    labels = [", ".join(n.split()[0] for n in p["members"]) for p in ps]
    ax.barh(list(y), [1]*len(ps), color=colors)
    for i, p in enumerate(ps):
        txt = "together" if p["satisfied"] else "split (no common choice)"
        ax.text(0.02, i, f"{labels[i]} — {txt}", va="center", fontsize=9, color="white")
    ax.set_yticks([]); ax.set_xticks([])
    ax.set_title(f"Partner Groups — {res['n_partners_ok']}/{len(ps)} kept together",
                 fontsize=13, fontweight="bold")
    ax.spines[:].set_visible(False)
    fig.tight_layout(); fig.savefig(path, dpi=150, facecolor="white"); plt.close(fig)


def chart_utilization(data, res, path):
    projects = data["projects"]
    items = sorted(res["opened"].items(), key=lambda kv: kv[1], reverse=True)
    names = [projects[p]["name"][:30] for p, _ in items]
    used = [n for _, n in items]
    cap = [projects[p]["quota"][1] for p, _ in items]
    y = range(len(items))
    fig, ax = plt.subplots(figsize=(9, 0.32*len(items)+1))
    ax.barh(y, cap, color="#ECF0F1", label="Capacity")
    ax.barh(y, used, color=C_BLUE, label="Assigned")
    ax.set_yticks(list(y)); ax.set_yticklabels(names, fontsize=8)
    ax.invert_yaxis(); ax.set_xlabel("Students"); ax.legend(loc="lower right")
    ax.set_title(f"Project Utilization — {res['n_opened']} projects opened",
                 fontsize=13, fontweight="bold")
    ax.spines[["top", "right"]].set_visible(False)
    fig.tight_layout(); fig.savefig(path, dpi=150, facecolor="white"); plt.close(fig)


def main():
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    data = load_data()
    res = solve(data, criterion="minimax_weighted")
    if not res["feasible"]:
        print("Baseline infeasible — aborting."); return

    print(f"Baseline: {res['rank_counts']}  satisfaction={res['satisfaction_pct']:.1f}%")
    save_excel(data, res, OUTPUT_DIR / "allocation_2026.xlsx")
    chart_rank(res, OUTPUT_DIR / "rank_distribution.png")
    chart_demand(data, OUTPUT_DIR / "demand_vs_capacity.png")
    chart_partners(res, OUTPUT_DIR / "partner_groups.png")
    chart_utilization(data, res, OUTPUT_DIR / "project_utilization.png")
    print("All outputs written to output/")


if __name__ == "__main__":
    main()
