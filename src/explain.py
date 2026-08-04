"""
explain.py
================================================================
Explainability for the allocation (project requirement §3.4).

Produces, for every student, a plain-language explanation of WHY they
received their assignment — understandable without optimisation theory:

  * If they got their 1st choice  → simply state it.
  * If they got a lower choice    → name the higher-ranked project(s)
    they missed and the binding reason (project full / closed below its
    minimum / taken by partner coupling).

Also identifies the binding constraints for the allocation as a whole
and writes a human-readable report (text + Excel).

Outputs:
    output/explanations.xlsx        per-student explanations
    output/explanation_report.txt   same content as plain text
================================================================
"""

from collections import defaultdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from spa_milp import load_data, load_student_tracks, solve

OUTPUT_DIR = Path(__file__).resolve().parent.parent / "output"
C_BLUE = "193764"


def build_explanations(data, res):
    """Return a list of {student, project, rank, explanation} dicts."""
    projects = data["projects"]
    students = data["students"]

    # occupancy of every project under this allocation
    occupancy = defaultdict(int)
    for a in res["assignments"]:
        if a["project"]:
            occupancy[a["project"]] += 1

    # which students are partnered (their placement is coupled)
    partnered = {sid for g in data["partner_groups"] if len(g) > 1 for sid in g}
    group_of = {}
    for gi, g in enumerate(data["partner_groups"]):
        for sid in g:
            group_of[sid] = gi

    explanations = []
    for a in res["assignments"]:
        s, p, rank = a["student"], a["project"], a["rank"]

        if p is None:
            explanations.append({
                "student": s["name"], "email": s["email"],
                "project": "NOT ASSIGNED", "rank": "-",
                "explanation": "No feasible project among the submitted choices.",
            })
            continue

        proj_name = projects[p]["name"]

        if rank == 1:
            reason = "Received their 1st choice."
        else:
            # explain each higher-ranked project that was missed
            missed = []
            for q in s["choices"]:
                if s["ranks"][q] < rank:
                    qmin, qmax = projects[q]["quota"]
                    if occupancy[q] >= qmax:
                        missed.append(f"\u201c{projects[q]['name'][:40]}\u201d was full "
                                      f"({qmax} place(s), all taken)")
                    elif occupancy[q] == 0:
                        missed.append(f"\u201c{projects[q]['name'][:40]}\u201d could not open "
                                      f"(needed at least {qmin}, too few applicants)")
                    else:
                        missed.append(f"\u201c{projects[q]['name'][:40]}\u201d was not "
                                      f"jointly feasible")
            why = "; ".join(missed) if missed else "higher choices were not feasible"
            reason = f"Received their choice #{rank}. Higher choices missed: {why}."

            # note partner coupling if relevant
            if s["id"] in partnered:
                mates = [o["name"] for o in students
                         if o["id"] != s["id"]
                         and group_of.get(o["id"]) == group_of.get(s["id"])]
                if mates:
                    reason += (f" Placement is coupled with partner(s) "
                               f"{', '.join(mates)}, which restricts the options to "
                               f"projects all partners ranked.")

        explanations.append({
            "student": s["name"], "email": s["email"],
            "project": proj_name, "rank": rank, "explanation": reason,
        })

    return explanations


def binding_constraints(data, res):
    """Summarise which constraints shaped the allocation overall."""
    projects = data["projects"]
    occupancy = defaultdict(int)
    for a in res["assignments"]:
        if a["project"]:
            occupancy[a["project"]] += 1

    full_projects = [p for p in res["opened"]
                     if occupancy[p] >= projects[p]["quota"][1]]
    closed_wanted = [p for p in res["forced_closed"]]

    lines = []
    lines.append(f"Worst rank enforced (z) : {res['worst_rank']}")
    lines.append(f"Projects opened          : {res['n_opened']}")
    lines.append(f"Projects at full capacity: {len(full_projects)} "
                 f"(these directly block lower-ranked demand)")
    lines.append(f"Projects force-closed    : {len(closed_wanted)} "
                 f"(fewer applicants than the minimum team size)")
    lines.append(f"Partner groups satisfied : {res['n_partners_ok']}/"
                 f"{len(res['partner_status'])}")
    return lines, full_projects


def save_excel(explanations, binding, path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Explanations"
    ws.append(["Student", "Email", "Assigned Project", "Rank", "Why"])
    for c in ws[1]:
        c.fill = PatternFill("solid", fgColor=C_BLUE)
        c.font = Font(color="FFFFFF", bold=True)
    for e in sorted(explanations, key=lambda e: (str(e["rank"]), e["student"])):
        ws.append([e["student"], e["email"], e["project"], e["rank"], e["explanation"]])
    for col, w in zip("ABCDE", [24, 30, 45, 6, 90]):
        ws.column_dimensions[col].width = w

    ws2 = wb.create_sheet("Binding constraints")
    ws2.append(["Allocation-level summary"])
    ws2["A1"].font = Font(bold=True)
    for line in binding:
        ws2.append([line])
    ws2.column_dimensions["A"].width = 70
    wb.save(path)


def main():
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    data = load_data()
    tracks = load_student_tracks()
    res = solve(data, criterion="minimax_weighted",
                track_lambda=0, student_tracks=tracks)
    if not res["feasible"]:
        print("Infeasible — aborting."); return

    explanations = build_explanations(data, res)
    binding, full_projects = binding_constraints(data, res)

    # text report
    txt = OUTPUT_DIR / "explanation_report.txt"
    with open(txt, "w", encoding="utf-8") as f:
        f.write("STUDENT–PROJECT ALLOCATION — EXPLANATION REPORT\n")
        f.write("=" * 60 + "\n\n")
        f.write("ALLOCATION-LEVEL BINDING CONSTRAINTS\n")
        f.write("-" * 60 + "\n")
        for line in binding:
            f.write("  " + line + "\n")
        f.write("\n\nPER-STUDENT EXPLANATIONS\n")
        f.write("-" * 60 + "\n")
        for e in sorted(explanations, key=lambda e: (str(e["rank"]), e["student"])):
            f.write(f"\n[{e['rank']}] {e['student']}  →  {e['project']}\n")
            f.write(f"     {e['explanation']}\n")
    print(f"Saved → {txt}")

    save_excel(explanations, binding, OUTPUT_DIR / "explanations.xlsx")
    print(f"Saved → {OUTPUT_DIR / 'explanations.xlsx'}")

    # console preview
    print("\nBinding constraints:")
    for line in binding:
        print("  " + line)
    print("\nSample explanations (non-first-choice students):")
    shown = 0
    for e in explanations:
        if e["rank"] not in (1, "-") and shown < 5:
            print(f"  • {e['student']} → #{e['rank']}: {e['explanation'][:100]}...")
            shown += 1


if __name__ == "__main__":
    main()
